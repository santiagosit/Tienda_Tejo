# Reporte de productos más vendidos
import openpyxl
from django.http import HttpResponse

from openpyxl.utils import get_column_letter


from app_inventario.models import Producto
from app_ventas.models import VentaDetalle
from django.db.models import  F


def productos_mas_vendidos(tipo_tiempo='mensual'):
    hoy = datetime.date.today()
    if tipo_tiempo == 'semanal':
        inicio_periodo = hoy - datetime.timedelta(days=7)
    else:  # mensual
        inicio_periodo = hoy.replace(day=1)

    # Productos más vendidos (ajustado al modelo)
    productos_vendidos = VentaDetalle.objects.filter(
        venta__fecha__range=[inicio_periodo, hoy]
    ).values('producto__nombre').annotate(
        total_vendido=Sum('cantidad')
    ).order_by('-total_vendido')

    return productos_vendidos

# Productos sin stock
def productos_sin_stock():
    return Producto.objects.filter(cantidad_stock=0)

# Productos que no se venden (últimos 30 días sin ventas)
def productos_no_vendidos():
    hoy = datetime.date.today()
    hace_30_dias = hoy - datetime.timedelta(days=30)

    productos_no_vendidos = Producto.objects.exclude(
        id__in=VentaDetalle.objects.filter(venta__fecha__gte=hace_30_dias).values('producto')
    )
    return productos_no_vendidos
def reporte_inventario(request):
    # Productos más vendidos (sumamos la cantidad vendida por producto)
    productos_mas_vendidos = VentaDetalle.objects.values('producto__nombre').annotate(total_vendido=Sum('cantidad')).order_by('-total_vendido')[:10]

    # Productos sin stock
    productos_sin_stock = Producto.objects.filter(cantidad_stock=0)

    # Productos que se agotan más rápido (podemos suponer que los que tienen bajo stock pero son muy vendidos)
    productos_agotandose = Producto.objects.filter(cantidad_stock__lt=F('stock_minimo')).order_by('cantidad_stock')

    # Productos que no se venden (aquellos que no tienen ventas asociadas)
    productos_no_vendidos = Producto.objects.exclude(ventadetalle__isnull=False)

    context = {
        'productos_mas_vendidos': productos_mas_vendidos,
        'productos_sin_stock': productos_sin_stock,
        'productos_agotandose': productos_agotandose,
        'productos_no_vendidos': productos_no_vendidos,
    }

    return render(request, 'reportes/reporte_inventario.html', context)


import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from django.http import HttpResponse
from django.db.models import Sum, F
from app_inventario.models import Producto
from app_ventas.models import VentaDetalle
from app_finanzas.models import Ingreso, Egreso
import datetime


def estilizar_encabezado(celda):
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    celda.alignment = Alignment(horizontal="center", vertical="center")
    celda.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                          bottom=Side(style='thin'))


def exportar_reporte_excel(request):
    hoy = datetime.date.today()

    # Creamos un nuevo libro de Excel
    wb = openpyxl.Workbook()

    # --- Hoja 1: Productos más vendidos ---
    ws1 = wb.active
    ws1.title = 'Productos más vendidos'

    # Encabezados
    ws1.append(['Producto', 'Total Vendido'])
    for celda in ws1[1]:
        estilizar_encabezado(celda)

    # Datos
    productos_mas_vendidos = VentaDetalle.objects.values('producto__nombre').annotate(
        total_vendido=Sum('cantidad')).order_by('-total_vendido')[:10]
    for producto in productos_mas_vendidos:
        ws1.append([producto['producto__nombre'], producto['total_vendido']])

    # Ajustar ancho de columnas
    for col in ws1.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws1.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Gráfico de barras para productos más vendidos
    chart1 = BarChart()
    data = Reference(ws1, min_col=2, min_row=1, max_row=len(productos_mas_vendidos) + 1)
    cats = Reference(ws1, min_col=1, min_row=2, max_row=len(productos_mas_vendidos) + 1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.title = "Productos más vendidos"
    chart1.y_axis.title = "Cantidad vendida"
    chart1.x_axis.title = "Producto"
    ws1.add_chart(chart1, "E5")

    # --- Hoja 2: Productos sin stock ---
    ws2 = wb.create_sheet('Productos sin stock')

    # Encabezados
    ws2.append(['Producto'])
    for celda in ws2[1]:
        estilizar_encabezado(celda)

    # Datos
    productos_sin_stock = Producto.objects.filter(cantidad_stock=0)
    for producto in productos_sin_stock:
        ws2.append([producto.nombre])

    # Ajustar ancho de columnas
    for col in ws2.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # --- Hoja 3: Productos que se agotan más rápido ---
    ws3 = wb.create_sheet('Productos agotándose')

    # Encabezados
    ws3.append(['Producto', 'Stock actual'])
    for celda in ws3[1]:
        estilizar_encabezado(celda)

    # Datos
    productos_agotandose = Producto.objects.filter(cantidad_stock__lt=F('stock_minimo')).order_by('cantidad_stock')
    for producto in productos_agotandose:
        ws3.append([producto.nombre, producto.cantidad_stock])

    # Gráfico de barras para productos que se agotan más rápido
    chart2 = BarChart()
    data = Reference(ws3, min_col=2, min_row=1, max_row=len(productos_agotandose) + 1)
    cats = Reference(ws3, min_col=1, min_row=2, max_row=len(productos_agotandose) + 1)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.title = "Productos con bajo stock"
    chart2.y_axis.title = "Stock actual"
    chart2.x_axis.title = "Producto"
    ws3.add_chart(chart2, "E5")

    # --- Hoja 4: Productos que no se venden ---
    ws4 = wb.create_sheet('Productos no vendidos')

    # Encabezados
    ws4.append(['Producto'])
    for celda in ws4[1]:
        estilizar_encabezado(celda)

    # Datos
    productos_no_vendidos = Producto.objects.exclude(ventadetalle__isnull=False)
    for producto in productos_no_vendidos:
        ws4.append([producto.nombre])


    # Configurar la respuesta HTTP para la descarga de Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reporte_inventario.xlsx'

    # Guardar el archivo en el response
    wb.save(response)
    return response


from django.shortcuts import render
from django.db.models import Sum
import datetime
from django.utils.timezone import make_aware

from app_finanzas.models import Egreso
from app_ventas.models import Venta  # Asegúrate de importar el modelo de ventas

from django.shortcuts import render
from django.db.models import Sum
import datetime
from django.utils.timezone import make_aware
from app_finanzas.models import Egreso
from app_ventas.models import Venta


from django.shortcuts import render
from django.db.models import Sum
from django.utils import timezone
from app_finanzas.models import Egreso
from app_ventas.models import Venta

def reporte_ingresos_egresos(request):
    hoy = timezone.now().date()  # Usamos timezone.now() en lugar de datetime.date.today()
    tipo_tiempo = request.GET.get('tipo_tiempo', 'mensual')

    # Definir el inicio del periodo según el tipo (mensual o semanal)
    if tipo_tiempo == 'semanal':
        inicio_periodo = hoy - timezone.timedelta(days=7)
    else:
        inicio_periodo = hoy.replace(day=1)

    # Ajustar las fechas a timezone-aware
    inicio_periodo = timezone.make_aware(datetime.datetime.combine(inicio_periodo, datetime.time.min))
    fin_periodo = timezone.make_aware(datetime.datetime.combine(hoy, datetime.time.max))

    # Calcular el total de ingresos basados en las ventas
    ingresos = Venta.objects.filter(fecha__range=[inicio_periodo, fin_periodo]).aggregate(total_ingresos=Sum('total'))['total_ingresos'] or 0

    # Calcular el total de egresos
    egresos = Egreso.objects.filter(fecha__range=[inicio_periodo, fin_periodo]).aggregate(total_egresos=Sum('monto'))['total_egresos'] or 0

    # Calcular el balance
    balance = ingresos - egresos

    context = {
        'ingresos': ingresos,
        'egresos': egresos,
        'balance': balance,
        'tipo_tiempo': tipo_tiempo,
    }

    return render(request, 'reportes/reporte_ingresos_egresos.html', context)



from django.shortcuts import render
from django.db.models import Sum, F, Q
from app_finanzas.models import Ingreso, Egreso
from app_ventas.models import Venta
from app_inventario.models import Producto
from app_pedidos.models import Pedido

def home(request):

    productos_bajo_stock_count = Producto.objects.filter(cantidad_stock__lt=F('stock_minimo')).count()
    productos_sin_stock_count = Producto.objects.filter(cantidad_stock=0).count()

    # Contador de pedidos pendientes
    pedidos_pendientes_count = Pedido.objects.filter(Q(estado='pedido') | Q(estado='en camino')).count()

    context = {


        'productos_bajo_stock_count': productos_bajo_stock_count,
        'productos_sin_stock_count': productos_sin_stock_count,
        'pedidos_pendientes_count': pedidos_pendientes_count,
    }

    return render(request, 'home.html', context)
